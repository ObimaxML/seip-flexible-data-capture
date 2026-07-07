from __future__ import annotations

import hmac
import os
import re
import shutil
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

try:
    from streamlit_js_eval import streamlit_js_eval
except Exception:
    streamlit_js_eval = None


APP_TITLE = "SEIP Flexible Data Capture"

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
ASSETS_DIR = BASE_DIR / "assets"

DEFAULT_TEMPLATE_PATH = DATA_DIR / "SEIP_Data_Collection_Template.xlsx"
CAPTURE_PATH = DATA_DIR / "SEIP_Captured_Data.xlsx"

AUTH_USERNAME = "seip_admin"
DEFAULT_AUTH_PASSWORD = "seip_admin123"

DATA_DIR.mkdir(exist_ok=True)
ASSETS_DIR.mkdir(exist_ok=True)

st.set_page_config(
    page_title=APP_TITLE,
    layout="wide",
    page_icon="📋",
)


# ============================================================
# MOBILE CSS
# ============================================================

def inject_mobile_css() -> None:
    st.markdown(
        """
        <style>
        @media (max-width: 768px) {

            .block-container {
                padding-left: 0.75rem !important;
                padding-right: 0.75rem !important;
                padding-top: 4.25rem !important;
                padding-bottom: 1rem !important;
            }

            header[data-testid="stHeader"] {
                height: 3.25rem !important;
            }

            section[data-testid="stSidebar"] {
                width: 88vw !important;
                z-index: 999999 !important;
            }

            section[data-testid="stSidebar"] > div:first-child {
                padding-top: 4.25rem !important;
            }

            section[data-testid="stSidebar"] img {
                max-width: 180px !important;
                height: auto !important;
                margin-top: 1.25rem !important;
            }

            h1 {
                font-size: 1.55rem !important;
                line-height: 1.2 !important;
            }

            h2, h3 {
                font-size: 1.25rem !important;
                line-height: 1.25 !important;
            }

            h4 {
                font-size: 1.05rem !important;
            }

            p, label, div, span {
                font-size: 0.95rem;
            }

            [data-testid="stImage"] img {
                max-width: 100% !important;
                height: auto !important;
            }

            [data-testid="stMetric"] {
                background: #f8fafc;
                border: 1px solid #e5e7eb;
                border-radius: 12px;
                padding: 0.75rem;
                margin-bottom: 0.5rem;
            }

            [data-testid="stMetricLabel"] {
                font-size: 0.75rem !important;
            }

            [data-testid="stMetricValue"] {
                font-size: 1.1rem !important;
            }

            div[data-testid="column"] {
                width: 100% !important;
                flex: 1 1 100% !important;
                min-width: 100% !important;
            }

            .stButton > button {
                width: 100% !important;
                min-height: 44px !important;
                border-radius: 10px !important;
                font-size: 0.95rem !important;
                margin-top: 0.25rem !important;
                margin-bottom: 0.25rem !important;
            }

            .stDownloadButton > button {
                width: 100% !important;
                min-height: 44px !important;
                border-radius: 10px !important;
                font-size: 0.95rem !important;
            }

            .stTextInput input,
            .stNumberInput input,
            .stDateInput input {
                min-height: 40px !important;
                height: 40px !important;
                font-size: 0.95rem !important;
            }

            .stSelectbox div[data-baseweb="select"] {
                min-height: 40px !important;
                height: 40px !important;
                font-size: 0.95rem !important;
            }

            textarea {
                font-size: 0.95rem !important;
                min-height: 80px !important;
            }

            [data-testid="stDataFrame"] {
                overflow-x: auto !important;
            }

            .element-container {
                margin-bottom: 0.45rem !important;
            }

            div[data-testid="stExpander"] {
                border-radius: 12px !important;
            }

            iframe {
                max-width: 100% !important;
            }
        }

        .stTextInput input,
        .stNumberInput input,
        .stDateInput input {
            min-height: 40px;
            height: 40px;
        }

        .stSelectbox div[data-baseweb="select"] {
            min-height: 40px;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


inject_mobile_css()


# ============================================================
# ID CONFIGURATION
# ============================================================

AREA_ID_CONFIG = {
    "job_seeker": {
        "prefix": "JS",
        "id_column": "job_seeker_id",
        "label": "Job Seeker",
    },
    "job_seekers": {
        "prefix": "JS",
        "id_column": "job_seeker_id",
        "label": "Job Seeker",
    },
    "informal_business": {
        "prefix": "IB",
        "id_column": "informal_business_id",
        "label": "Informal Business",
    },
    "informal_businesses": {
        "prefix": "IB",
        "id_column": "informal_business_id",
        "label": "Informal Business",
    },
    "business": {
        "prefix": "BZ",
        "id_column": "business_id",
        "label": "Business",
    },
    "businesses": {
        "prefix": "BZ",
        "id_column": "business_id",
        "label": "Business",
    },
    "training_provider": {
        "prefix": "TP",
        "id_column": "training_provider_id",
        "label": "Training Provider",
    },
    "training_providers": {
        "prefix": "TP",
        "id_column": "training_provider_id",
        "label": "Training Provider",
    },
}

DEFAULT_ID_CONFIG = {
    "prefix": "SEIP",
    "id_column": "record_id",
    "label": "Record",
}

ID_WIDTH = 3

ID_COLUMNS = [
    "record_id",
    "respondent_id",
    "job_seeker_id",
    "informal_business_id",
    "business_id",
    "training_provider_id",
    "person_id",
    "seeker_id",
    "participant_id",
]

GPS_COLUMNS = [
    "gps_capture_required",
    "gps_capture_method",
    "gps_latitude",
    "gps_longitude",
    "gps_coordinates",
    "gps_accuracy_meters",
    "gps_captured_at",
]

SYSTEM_COLUMNS = [
    "created_at",
    "updated_at",
    "captured_by",
    "updated_by",
]


# ============================================================
# AUTH
# ============================================================

def get_auth_password() -> str:
    try:
        return str(st.secrets.get("SEIP_ADMIN_PASSWORD", DEFAULT_AUTH_PASSWORD))
    except Exception:
        return os.getenv("SEIP_ADMIN_PASSWORD", DEFAULT_AUTH_PASSWORD)


def show_branding() -> None:
    banner = ASSETS_DIR / "seip_banner.svg"
    logo = ASSETS_DIR / "seip_logo.svg"

    if banner.exists():
        st.image(str(banner), use_container_width=True)
    elif logo.exists():
        st.image(str(logo), width=220)
    else:
        st.warning(
            "Logo/banner not found. Add `assets/seip_logo.svg` and "
            "`assets/seip_banner.svg`."
        )


def login_screen() -> None:
    show_branding()

    st.title("SEIP Data Capture Login")
    st.caption("Sign in to access the SEIP capture app.")

    with st.form("login_form"):
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        submitted = st.form_submit_button("Login", type="primary")

    if submitted:
        valid_user = hmac.compare_digest(username.strip(), AUTH_USERNAME)
        valid_password = hmac.compare_digest(password, get_auth_password())

        if valid_user and valid_password:
            st.session_state["authenticated"] = True
            st.session_state["username"] = AUTH_USERNAME
            st.rerun()
        else:
            st.error("Invalid username or password.")


def require_login() -> None:
    if not st.session_state.get("authenticated", False):
        login_screen()
        st.stop()


def logout_button() -> None:
    if st.sidebar.button("Logout"):
        st.session_state.clear()
        st.rerun()


# ============================================================
# HELPERS
# ============================================================

def normalise_label(name: str) -> str:
    return str(name).replace("_", " ").strip().title()


def normalise_key(name: str) -> str:
    return (
        str(name)
        .strip()
        .lower()
        .replace(" ", "_")
        .replace("-", "_")
        .replace("/", "_")
    )


def get_area_id_config(area_key: str, capture_sheet: str) -> Dict[str, str]:
    area_key_norm = normalise_key(area_key)
    capture_sheet_norm = normalise_key(capture_sheet)

    candidates = [area_key_norm, capture_sheet_norm]

    for candidate in candidates:
        if candidate in AREA_ID_CONFIG:
            return AREA_ID_CONFIG[candidate]

    for candidate in candidates:
        if "job" in candidate and "seeker" in candidate:
            return AREA_ID_CONFIG["job_seeker"]

        if "informal" in candidate and "business" in candidate:
            return AREA_ID_CONFIG["informal_business"]

        if "training" in candidate and "provider" in candidate:
            return AREA_ID_CONFIG["training_provider"]

        if "business" in candidate:
            return AREA_ID_CONFIG["business"]

    return DEFAULT_ID_CONFIG


def ensure_capture_workbook(source_template: Path) -> Path:
    if not CAPTURE_PATH.exists():
        shutil.copy(source_template, CAPTURE_PATH)

    return CAPTURE_PATH


@st.cache_data(show_spinner=False)
def load_excel_sheets(path_str: str) -> Dict[str, pd.DataFrame]:
    return pd.read_excel(path_str, sheet_name=None, engine="openpyxl")


def read_workbook(path: Path) -> Dict[str, pd.DataFrame]:
    return load_excel_sheets(str(path))


def save_uploaded_template(uploaded_file) -> Path:
    upload_path = DATA_DIR / "uploaded_template.xlsx"

    with open(upload_path, "wb") as f:
        f.write(uploaded_file.getbuffer())

    load_excel_sheets.clear()
    return upload_path


def is_empty_value(value: Any) -> bool:
    if value is None:
        return True

    try:
        if pd.isna(value):
            return True
    except Exception:
        pass

    return str(value).strip() == ""


# ============================================================
# SOUTH AFRICAN ID VALIDATION
# ============================================================

def is_sa_id_number(value: Any) -> bool:
    id_number = re.sub(r"\D", "", str(value or ""))

    if len(id_number) != 13:
        return False

    try:
        yy = int(id_number[0:2])
        mm = int(id_number[2:4])
        dd = int(id_number[4:6])
    except Exception:
        return False

    current_yy = int(str(datetime.now().year)[-2:])
    century = 2000 if yy <= current_yy else 1900

    try:
        date(century + yy, mm, dd)
    except ValueError:
        return False

    try:
        total_odd = sum(int(id_number[i]) for i in range(0, 12, 2))
        even_digits = "".join(id_number[i] for i in range(1, 12, 2))
        doubled_even = str(int(even_digits) * 2)
        total_even = sum(int(d) for d in doubled_even)
        check_digit = (10 - ((total_odd + total_even) % 10)) % 10
        return check_digit == int(id_number[-1])
    except Exception:
        return False


def looks_like_id_number_field(field_name: str) -> bool:
    field = normalise_key(field_name)

    return field in {
        "id_number",
        "sa_id_number",
        "south_african_id",
        "south_african_id_number",
        "rsa_id",
        "rsa_id_number",
        "national_id",
        "identity_number",
    }


# ============================================================
# TEMPLATE DETECTION
# ============================================================

def detect_capture_sheets(
    sheets: Dict[str, pd.DataFrame],
) -> Dict[str, Dict[str, Any]]:
    areas: Dict[str, Dict[str, Any]] = {}
    all_names = set(sheets.keys())

    for sheet_name, df in sheets.items():
        if not sheet_name.endswith("_fields"):
            continue

        if "field_name" not in df.columns:
            continue

        base_name = sheet_name[: -len("_fields")]
        capture_sheet = base_name if base_name in all_names else None

        areas[base_name] = {
            "label": normalise_label(base_name),
            "fields_sheet": sheet_name,
            "capture_sheet": capture_sheet,
            "fields": df.dropna(how="all").copy(),
        }

    if not areas:
        skip = {
            "README",
            "reference_values",
            "validation_rules",
        }

        for sheet_name, df in sheets.items():
            if sheet_name in skip or df.empty:
                continue

            field_rows = []

            for col in df.columns:
                col_name = str(col).strip()

                if col_name in ID_COLUMNS:
                    continue

                if col_name in GPS_COLUMNS:
                    continue

                if col_name in SYSTEM_COLUMNS:
                    continue

                field_rows.append(
                    {
                        "field_name": col_name,
                        "data_type": "TEXT",
                        "required": "No",
                        "example": "",
                        "description": "",
                        "validation_rule": "",
                    }
                )

            areas[sheet_name] = {
                "label": normalise_label(sheet_name),
                "fields_sheet": None,
                "capture_sheet": sheet_name,
                "fields": pd.DataFrame(field_rows),
            }

    return areas


def reference_options(sheets: Dict[str, pd.DataFrame]) -> Dict[str, List[Any]]:
    ref_map: Dict[str, List[Any]] = {}
    ref_df = sheets.get("reference_values")

    if ref_df is None or ref_df.empty:
        return ref_map

    required_cols = {"reference_type", "code"}

    if not required_cols.issubset(ref_df.columns):
        return ref_map

    for ref_type, group in ref_df.dropna(
        subset=["reference_type", "code"],
    ).groupby("reference_type"):
        ref_map[str(ref_type)] = [
            x for x in group["code"].tolist() if pd.notna(x)
        ]

    return ref_map


def match_options_for_field(
    field_name: str,
    ref_map: Dict[str, List[Any]],
) -> List[Any] | None:
    if field_name in ref_map:
        return ref_map[field_name]

    mappings = {
        "township": "township",
        "gender": "gender",
        "digital_literacy_level": "digital_literacy_level",
        "transport_mode": "transport_mode",
        "sector_code": "sector_code",
        "preferred_sector": "sector_code",
        "previous_sector": "sector_code",
        "business_sector": "sector_code",
        "training_sector": "sector_code",
        "ward": "ward",
        "ward_number": "ward",
        "highest_qualification": "highest_qualification",
        "qualification": "highest_qualification",
    }

    ref_key = mappings.get(field_name)

    if ref_key and ref_key in ref_map:
        return ref_map[ref_key]

    return None


# ============================================================
# FORM RENDERING
# ============================================================

def clean_example(value: Any) -> Any:
    if pd.isna(value):
        return None

    return value


def infer_widget_type(field_name: str, data_type: str, example: Any) -> str:
    field = normalise_key(field_name)
    dtype = str(data_type or "").upper()

    if looks_like_id_number_field(field_name):
        return "id_number"

    if "BOOLEAN" in dtype or isinstance(example, bool):
        return "boolean"

    if "DATE" in dtype or isinstance(example, (datetime, date, pd.Timestamp)):
        return "date"

    if any(x in dtype for x in ["INT", "BIGINT", "SMALLINT"]):
        return "integer"

    if any(x in dtype for x in ["DECIMAL", "NUMERIC", "FLOAT", "DOUBLE", "REAL"]):
        return "decimal"

    textarea_keywords = [
        "comment",
        "comments",
        "note",
        "notes",
        "description",
        "remarks",
        "motivation",
        "challenge",
        "challenges",
        "other_specify",
    ]

    if any(keyword in field for keyword in textarea_keywords):
        return "textarea"

    return "text"


def reset_form(area_key: str) -> None:
    st.session_state[f"reset_version_{area_key}"] = (
        st.session_state.get(f"reset_version_{area_key}", 0) + 1
    )

    keys_to_delete = [
        key
        for key in st.session_state.keys()
        if key.startswith(f"{area_key}_gps_detected_")
        or key.startswith(f"{area_key}_browser_gps_")
    ]

    for key in keys_to_delete:
        del st.session_state[key]

    st.rerun()


def get_form_version(area_key: str) -> int:
    return int(st.session_state.get(f"reset_version_{area_key}", 0))


def build_form(
    area_key: str,
    fields_df: pd.DataFrame,
    ref_map: Dict[str, List[Any]],
    form_version: int,
) -> Tuple[Dict[str, Any], List[str]]:
    values: Dict[str, Any] = {}
    validation_errors: List[str] = []

    for _, row in fields_df.iterrows():
        field_name = str(row.get("field_name", "")).strip()

        if not field_name or field_name.lower() == "nan":
            continue

        if field_name in ID_COLUMNS:
            continue

        if field_name in GPS_COLUMNS:
            continue

        if field_name in SYSTEM_COLUMNS:
            continue

        data_type = str(row.get("data_type", "TEXT"))

        description = (
            ""
            if pd.isna(row.get("description"))
            else str(row.get("description"))
        )

        validation_rule = (
            ""
            if pd.isna(row.get("validation_rule"))
            else str(row.get("validation_rule"))
        )

        example = clean_example(row.get("example"))
        widget_type = infer_widget_type(field_name, data_type, example)
        options = match_options_for_field(field_name, ref_map)

        label = normalise_label(field_name)

        if looks_like_id_number_field(field_name):
            label = f"{label} - South African ID"

        help_text = "\n".join(
            [
                x
                for x in [
                    description,
                    f"Validation: {validation_rule}" if validation_rule else "",
                ]
                if x
            ]
        )

        key = f"{area_key}_{field_name}_{form_version}"

        if options:
            choices = [""] + [str(x) for x in options]

            value = st.selectbox(
                label,
                choices,
                index=0,
                help=help_text,
                key=key,
            )
            value = None if value == "" else value

        elif widget_type == "boolean":
            value = st.checkbox(
                label,
                value=False,
                help=help_text,
                key=key,
            )

        elif widget_type == "date":
            value = st.date_input(
                label,
                value=None,
                help=help_text,
                key=key,
            )

        elif widget_type == "integer":
            value = st.number_input(
                label,
                value=None,
                step=1,
                format="%d",
                help=help_text,
                key=key,
            )

            value = None if value is None else int(value)

        elif widget_type == "decimal":
            value = st.number_input(
                label,
                value=None,
                step=0.000001,
                format="%.6f",
                help=help_text,
                key=key,
            )

            value = None if value is None else float(value)

        elif widget_type == "textarea":
            value = st.text_area(
                label,
                value="",
                help=help_text,
                key=key,
                height=90,
            )

        elif widget_type == "id_number":
            value = st.text_input(
                label,
                value="",
                help="Optional. If entered, it must be a valid 13-digit South African ID number.",
                key=key,
                max_chars=13,
            )

            cleaned_id = re.sub(r"\D", "", value)

            if value and cleaned_id != value:
                st.warning("ID number must contain digits only.")

            value = cleaned_id

            if value:
                if is_sa_id_number(value):
                    st.success("Valid South African ID number.")
                else:
                    st.error("Invalid South African ID number.")
                    validation_errors.append("ID number is not a valid South African ID.")

        else:
            value = st.text_input(
                label,
                value="",
                help=help_text,
                key=key,
            )

        if isinstance(value, date):
            values[field_name] = value.isoformat()
        else:
            values[field_name] = value

    return values, validation_errors


# ============================================================
# ID GENERATION
# ============================================================

def next_dynamic_id_for_area(
    workbook_path: Path,
    sheet_name: str,
    prefix: str,
    width: int = ID_WIDTH,
) -> str:
    highest = 0
    pattern = re.compile(rf"^{re.escape(prefix)}(\d+)$", re.IGNORECASE)

    if not workbook_path.exists():
        return f"{prefix}{1:0{width}d}"

    try:
        df = pd.read_excel(
            workbook_path,
            sheet_name=sheet_name,
            engine="openpyxl",
            dtype=str,
        )

        for column in df.columns:
            column_name = str(column).strip()

            if column_name not in ID_COLUMNS:
                continue

            for raw_value in df[column].dropna().astype(str):
                match = pattern.match(raw_value.strip())

                if match:
                    highest = max(highest, int(match.group(1)))

    except Exception:
        pass

    return f"{prefix}{highest + 1:0{width}d}"


# ============================================================
# GPS CAPTURE
# ============================================================

def get_browser_gps(component_key: str) -> Dict[str, Any] | None:
    if streamlit_js_eval is None:
        st.error(
            "Automatic GPS requires `streamlit-js-eval`. "
            "Install it with: pip install streamlit-js-eval"
        )
        return None

    gps = streamlit_js_eval(
        js_expressions="""
        new Promise((resolve) => {
            if (!navigator.geolocation) {
                resolve({
                    ok: false,
                    error: "Geolocation is not supported by this browser."
                });
            } else {
                navigator.geolocation.getCurrentPosition(
                    (position) => {
                        resolve({
                            ok: true,
                            latitude: position.coords.latitude,
                            longitude: position.coords.longitude,
                            accuracy: position.coords.accuracy,
                            timestamp: position.timestamp
                        });
                    },
                    (error) => {
                        let message = "Unknown GPS error";

                        if (error.code === 1) {
                            message = "Location permission was denied. Please allow location access in the browser.";
                        } else if (error.code === 2) {
                            message = "Location unavailable. Turn on device location/GPS.";
                        } else if (error.code === 3) {
                            message = "GPS request timed out. Try again.";
                        }

                        resolve({
                            ok: false,
                            error: message,
                            code: error.code
                        });
                    },
                    {
                        enableHighAccuracy: true,
                        timeout: 20000,
                        maximumAge: 0
                    }
                );
            }
        })
        """,
        key=component_key,
    )

    return gps


def build_gps_capture(
    area_key: str,
    form_version: int,
) -> Dict[str, Any]:
    st.markdown("#### GPS location capture")
    st.caption(
        "GPS coordinates are detected automatically from the browser/device. "
        "No manual latitude or longitude entry is required."
    )

    gps_state_key = f"{area_key}_gps_detected_{form_version}"

    gps_capture_required = st.checkbox(
        "Detect GPS coordinates for this record",
        value=False,
        key=f"{area_key}_gps_capture_required_{form_version}",
    )

    if not gps_capture_required:
        st.info("GPS Coordinates: Not captured")
        return {
            "gps_capture_required": False,
            "gps_capture_method": "not_captured",
            "gps_latitude": None,
            "gps_longitude": None,
            "gps_coordinates": None,
            "gps_accuracy_meters": None,
            "gps_captured_at": None,
        }

    st.warning(
        "Click `Detect GPS automatically`, then allow location access in your browser. "
        "GPS works best on a phone and requires localhost or HTTPS."
    )

    detect_button = st.button(
        "Detect GPS automatically",
        use_container_width=True,
        key=f"{area_key}_detect_gps_btn_{form_version}",
    )

    if detect_button:
        with st.spinner("Requesting GPS location from browser..."):
            gps_result = get_browser_gps(
                component_key=f"{area_key}_browser_gps_{form_version}_{datetime.now().timestamp()}",
            )

        if gps_result is not None:
            st.session_state[gps_state_key] = gps_result
            st.rerun()

    gps_result = st.session_state.get(gps_state_key)

    gps_record = {
        "gps_capture_required": True,
        "gps_capture_method": "not_captured",
        "gps_latitude": None,
        "gps_longitude": None,
        "gps_coordinates": None,
        "gps_accuracy_meters": None,
        "gps_captured_at": None,
    }

    if gps_result and isinstance(gps_result, dict):
        if gps_result.get("ok"):
            latitude = float(gps_result.get("latitude"))
            longitude = float(gps_result.get("longitude"))
            accuracy = gps_result.get("accuracy", None)
            gps_coordinates = f"{latitude:.6f}, {longitude:.6f}"

            gps_record.update(
                {
                    "gps_capture_required": True,
                    "gps_capture_method": "browser_geolocation",
                    "gps_latitude": latitude,
                    "gps_longitude": longitude,
                    "gps_coordinates": gps_coordinates,
                    "gps_accuracy_meters": float(accuracy) if accuracy is not None else None,
                    "gps_captured_at": datetime.now().isoformat(timespec="seconds"),
                }
            )

            st.success("GPS detected successfully.")
            st.markdown("##### Captured GPS Coordinates")
            st.code(gps_coordinates)

            if accuracy is not None:
                st.caption(f"Approximate accuracy: {float(accuracy):.1f} metres")

            map_df = pd.DataFrame(
                {
                    "lat": [latitude],
                    "lon": [longitude],
                }
            )

            st.map(
                map_df,
                latitude="lat",
                longitude="lon",
                zoom=15,
            )

        else:
            st.error(
                "Could not detect GPS automatically: "
                + str(gps_result.get("error", "Unknown browser GPS error"))
            )
            st.markdown("##### Captured GPS Coordinates")
            st.code("Not captured yet")

    else:
        st.info(
            "GPS has not been captured yet. Click `Detect GPS automatically` "
            "and allow location access in the browser."
        )
        st.markdown("##### Captured GPS Coordinates")
        st.code("Not captured yet")

    return gps_record


# ============================================================
# EXCEL SAVE / READ / DELETE / RESET / UPDATE
# ============================================================

def append_record(
    workbook_path: Path,
    sheet_name: str,
    record: Dict[str, Any],
) -> None:
    wb = load_workbook(workbook_path)

    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(list(record.keys()))
    else:
        ws = wb[sheet_name]

    headers = [cell.value for cell in ws[1]] if ws.max_row >= 1 else []
    headers = [h for h in headers if h is not None]

    for field_name in record.keys():
        if field_name not in headers:
            ws.cell(
                row=1,
                column=len(headers) + 1,
                value=field_name,
            )
            headers.append(field_name)

    row_values = [record.get(header, None) for header in headers]

    ws.append(row_values)
    wb.save(workbook_path)
    load_excel_sheets.clear()


def overwrite_sheet_records(
    workbook_path: Path,
    sheet_name: str,
    df: pd.DataFrame,
) -> None:
    wb = load_workbook(workbook_path)

    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)

    ws = wb[sheet_name]
    ws.delete_rows(1, ws.max_row)

    ws.append(list(df.columns))

    for row in df.itertuples(index=False, name=None):
        clean_row = []

        for value in row:
            if is_empty_value(value):
                clean_row.append(None)
            else:
                clean_row.append(value)

        ws.append(clean_row)

    wb.save(workbook_path)
    load_excel_sheets.clear()


def reset_sheet_records(
    workbook_path: Path,
    sheet_name: str,
    headers: List[str],
) -> None:
    wb = load_workbook(workbook_path)

    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
    else:
        ws = wb[sheet_name]

    ws.delete_rows(1, ws.max_row)
    ws.append(headers)

    wb.save(workbook_path)
    load_excel_sheets.clear()


def dataframe_for_sheet(
    path: Path,
    sheet_name: str,
) -> pd.DataFrame:
    try:
        return pd.read_excel(
            path,
            sheet_name=sheet_name,
            engine="openpyxl",
        )
    except Exception:
        return pd.DataFrame()


def file_download_button(
    path: Path,
    label: str,
    file_name: str,
    mime: str,
) -> None:
    if path.exists():
        st.download_button(
            label,
            data=path.read_bytes(),
            file_name=file_name,
            mime=mime,
        )


def validate_inline_dataframe(df: pd.DataFrame) -> List[str]:
    errors: List[str] = []

    for column in df.columns:
        if looks_like_id_number_field(column):
            for idx, value in df[column].items():
                if not is_empty_value(value):
                    cleaned_id = re.sub(r"\D", "", str(value))

                    if not is_sa_id_number(cleaned_id):
                        errors.append(
                            f"Row {idx}: `{column}` is not a valid South African ID number."
                        )

    return errors


# ============================================================
# MAIN APP
# ============================================================

require_login()

show_branding()

st.title(APP_TITLE)
st.caption("Capture SEIP survey data directly from the Excel template structure.")

with st.sidebar:
    logo = ASSETS_DIR / "seip_logo.svg"

    if logo.exists():
        st.image(str(logo), use_container_width=True)
    else:
        st.info("Sidebar logo missing: add `assets/seip_logo.svg`.")

    st.success(f"Logged in as {st.session_state.get('username', AUTH_USERNAME)}")
    logout_button()

    st.header("Workbook")

    uploaded = st.file_uploader(
        "Upload a different Excel template",
        type=["xlsx"],
    )

    if uploaded is not None:
        template_path = save_uploaded_template(uploaded)
        st.success("Uploaded template loaded for this session.")
    else:
        template_path = DEFAULT_TEMPLATE_PATH

    if not template_path.exists():
        st.error(
            "Template file not found. Please upload an Excel template or place it in "
            "`data/SEIP_Data_Collection_Template.xlsx`."
        )
        st.stop()

    capture_path = ensure_capture_workbook(template_path)

    st.write(f"**Capture file:** `{capture_path.name}`")

    file_download_button(
        capture_path,
        "Download captured Excel workbook",
        "SEIP_Captured_Data.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

sheets = read_workbook(capture_path)
areas = detect_capture_sheets(sheets)
ref_map = reference_options(sheets)

if not areas:
    st.warning(
        "No capture sheets were detected. Check that your workbook has sheets "
        "with headers or *_fields definition sheets."
    )
    st.stop()

area_labels = {
    meta["label"]: key
    for key, meta in areas.items()
}

selected_label = st.sidebar.radio(
    "Capture area",
    list(area_labels.keys()),
)

area_key = area_labels[selected_label]
area = areas[area_key]
capture_sheet = area.get("capture_sheet") or area_key
fields_df = area["fields"]

area_id_config = get_area_id_config(area_key, capture_sheet)

id_prefix = area_id_config["prefix"]
id_column = area_id_config["id_column"]
id_label = area_id_config["label"]

current_df = dataframe_for_sheet(
    path=capture_path,
    sheet_name=capture_sheet,
)

next_id = next_dynamic_id_for_area(
    workbook_path=capture_path,
    sheet_name=capture_sheet,
    prefix=id_prefix,
)

form_version = get_form_version(area_key)

st.subheader(selected_label)

metric_cols = st.columns(4)

metric_cols[0].metric(f"Next {id_label} ID", next_id)
metric_cols[1].metric("ID Prefix", id_prefix)
metric_cols[2].metric("ID Column", id_column)
metric_cols[3].metric("Capture Sheet", capture_sheet)

st.write("Fields are generated from the template. Records can be saved with minimal information.")

left, right = st.columns([1.2, 1])

with left:
    active_id = next_id

    st.markdown(f"#### New {id_label} ID: `{active_id}`")

    values, validation_errors = build_form(
        area_key=area_key,
        fields_df=fields_df,
        ref_map=ref_map,
        form_version=form_version,
    )

    gps_values = build_gps_capture(
        area_key=area_key,
        form_version=form_version,
    )

    button_col1, button_col2 = st.columns(2)

    with button_col1:
        submitted = st.button(
            "Save new record",
            type="primary",
            use_container_width=True,
        )

    with button_col2:
        reset_clicked = st.button(
            "Reset form fields",
            use_container_width=True,
        )

    if reset_clicked:
        reset_form(area_key)

    if submitted:
        gps_required = bool(gps_values.get("gps_capture_required"))
        gps_coordinates = gps_values.get("gps_coordinates")

        if validation_errors:
            st.error("Please fix: " + "; ".join(validation_errors))

        elif gps_required and not gps_coordinates:
            st.error(
                "GPS detection was selected, but coordinates were not captured. "
                "Click `Detect GPS automatically` and allow location access."
            )

        else:
            record = {
                id_column: active_id,
                **values,
                **gps_values,
                "created_at": datetime.now().isoformat(timespec="seconds"),
                "captured_by": st.session_state.get("username", AUTH_USERNAME),
            }

            append_record(
                workbook_path=capture_path,
                sheet_name=capture_sheet,
                record=record,
            )

            st.success(f"Record `{active_id}` saved to `{capture_sheet}`.")
            reset_form(area_key)

with right:
    st.markdown("#### Template notes")

    notes_cols = [
        c
        for c in [
            "field_name",
            "required",
            "data_type",
            "description",
            "validation_rule",
        ]
        if c in fields_df.columns
    ]

    if notes_cols:
        st.dataframe(
            fields_df[notes_cols],
            use_container_width=True,
            hide_index=True,
            height=420,
        )
    else:
        st.info("No template notes found.")


# ============================================================
# INLINE EDITING SECTION
# ============================================================

st.markdown("---")
st.markdown(f"### Captured data: `{capture_sheet}`")

current_df = dataframe_for_sheet(
    path=capture_path,
    sheet_name=capture_sheet,
)

if current_df.empty:
    st.info("No records found yet for this sheet.")
else:
    st.caption(
        "Edit existing records directly in the table below, then click "
        "`Save inline edits` to write the changes back to the Excel workbook."
    )

    disabled_columns = [
        col
        for col in current_df.columns
        if col in ID_COLUMNS
        or col in [
            "created_at",
            "captured_by",
        ]
    ]

    edited_df = st.data_editor(
        current_df,
        use_container_width=True,
        hide_index=False,
        num_rows="fixed",
        disabled=disabled_columns,
        key=f"inline_editor_{capture_sheet}_{form_version}",
    )

    edit_btn_col1, edit_btn_col2, edit_btn_col3 = st.columns([1, 1, 2])

    with edit_btn_col1:
        if st.button(
            "Save inline edits",
            type="primary",
            use_container_width=True,
            key=f"save_inline_edits_{capture_sheet}",
        ):
            inline_errors = validate_inline_dataframe(edited_df)

            if inline_errors:
                st.error("Please fix before saving:")
                for error in inline_errors:
                    st.write(f"- {error}")
            else:
                edited_df = edited_df.copy()
                edited_df["updated_at"] = datetime.now().isoformat(timespec="seconds")
                edited_df["updated_by"] = st.session_state.get("username", AUTH_USERNAME)

                overwrite_sheet_records(
                    workbook_path=capture_path,
                    sheet_name=capture_sheet,
                    df=edited_df,
                )

                st.success(f"Inline edits saved to `{capture_sheet}`.")
                st.rerun()

    with edit_btn_col2:
        csv_bytes = edited_df.to_csv(index=False).encode("utf-8")

        st.download_button(
            "Download CSV",
            csv_bytes,
            file_name=f"{capture_sheet}.csv",
            mime="text/csv",
            use_container_width=True,
        )

    st.markdown("#### Delete or reset records")

    delete_col, reset_col, spacer_col = st.columns([1, 1, 2])

    with delete_col:
        row_count = len(current_df)

        delete_index = st.number_input(
            "Delete row number",
            min_value=0,
            max_value=max(row_count - 1, 0),
            value=0,
            step=1,
            key=f"delete_index_{capture_sheet}",
        )

        if st.button(
            "Delete selected row",
            use_container_width=True,
            key=f"delete_selected_row_{capture_sheet}",
        ):
            new_df = current_df.drop(
                current_df.index[int(delete_index)]
            ).reset_index(drop=True)

            overwrite_sheet_records(
                workbook_path=capture_path,
                sheet_name=capture_sheet,
                df=new_df,
            )

            st.success(f"Deleted row {delete_index} from `{capture_sheet}`.")
            st.rerun()

    with reset_col:
        confirm_reset = st.checkbox(
            "Confirm reset sheet",
            key=f"confirm_reset_{capture_sheet}",
        )

        if st.button(
            "Reset saved records",
            disabled=not confirm_reset,
            use_container_width=True,
            key=f"reset_saved_records_{capture_sheet}",
        ):
            headers_to_keep = list(current_df.columns)

            reset_sheet_records(
                workbook_path=capture_path,
                sheet_name=capture_sheet,
                headers=headers_to_keep,
            )

            st.success(
                f"All saved records were cleared from `{capture_sheet}`. Headers were kept."
            )
            st.rerun()


with st.expander("Detected sheets, reference dropdowns and ID rules"):
    st.write(
        "Detected capture areas:",
        [meta["label"] for meta in areas.values()],
    )

    st.write("Reference lists:", sorted(ref_map.keys()))

    st.write(
        "ID rules:",
        {
            "Job Seeker": "JS001, JS002, JS003...",
            "Informal Business": "IB001, IB002, IB003...",
            "Business": "BZ001, BZ002, BZ003...",
            "Training Provider": "TP001, TP002, TP003...",
        },
    )

    st.write("GPS columns saved:", GPS_COLUMNS)